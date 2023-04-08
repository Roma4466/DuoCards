import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
from datetime import datetime
from analitics import DataAnalyzer

data = load_workbook('Copy.xlsx')
sheet = data.active

previous = 0
current = 0

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[1].value is None:
        break

    for i in range(4, min(len(row), 10)):
        if row[i].value is None:
            break
        current += row[i].value

    if current != 0 and previous != 0:
        print("Previous:", row[1].value)
        actual = current - previous
        print("Actual:", actual)
        row[1].value = actual

    previous = current
    current = 0

# data.save("Copy.xlsx")