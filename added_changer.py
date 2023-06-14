from openpyxl import load_workbook

from data_analyzer import draw_graphics
from data_list import DataHolder
from data_printer import prints_lists_values

data = load_workbook('data.xlsx')
sheet = data.active

previous = 0
current = 0
data_sum = 0

added_copy = {}
difference = {}

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[1].value is None:
        break

    data_sum += row[1].value

    for i in range(4, min(len(row), 10)):
        if row[i].value is None:
            break
        current += row[i].value

    if current != 0 and previous != 0:
        actual = current - previous
        if data_sum != actual:
            print(row[0].value.strftime("%d/%m"))
            print("Previous sum:", data_sum)
            print("Actual sum:", actual)
            print("Difference:", data_sum - actual)
            print("Was:", row[1].value)
            row[1].value -= (data_sum - actual)
            if data_sum == 8888:
                print()
            added_copy[row[0].value] = actual
            difference[row[0].value] = (data_sum - actual)
            print("New value:", row[1].value)
            print()

    if current != 0:
        previous = current
        data_sum = 0
    current = 0

data.save("Copy.xlsx")

calendar = list(difference.keys())
added_copy_list = list(added_copy.values())
difference_list = list(difference.values())

ratio_list = [added_copy_list[i] / difference_list[i] for i in range(len(difference))]
data_printer = DataHolder(
    calendar,
    "Copy",
    (ratio_list, "ratio_list")
)
prints_lists_values(data_printer)
try:
    draw_graphics(data_printer, float(input("Write down pause time: ")))
except:
    draw_graphics(data_printer, 0.01)

#%%
