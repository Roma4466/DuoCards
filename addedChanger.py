from openpyxl import load_workbook

data = load_workbook('data.xlsx')
sheet = data.active

previous = 0
current = 0
data_sum = 0

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    if row[1].value is None:
        break

    data_sum += row[1].value

    for i in range(4, min(len(row), 10)):
        if row[i].value is None:
            break
        current += row[i].value

    if current != 0 and previous != 0:
        actual = current - previous
        if data_sum != actual:
            print(row[0].value.strftime("%d/%m"))
            print("Previous sum:", data_sum)
            print("Actual sum:", actual)
            print("Difference:", data_sum - actual)
            row[1].value -= (data_sum - actual)
            print("New value:", row[1].value)
            print()

    if current != 0:
        previous = current
        data_sum = 0
    current = 0

data.save("Copy.xlsx")
