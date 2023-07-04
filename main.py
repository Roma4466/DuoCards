"""
This file reads data.xlxs,
puts data into dataHolder DTO
and starts data analytic
"""

from openpyxl import load_workbook
from data_analyzer import draw_graphics
from data_list import *
from data_printer import prints_lists_values
from sklearn.linear_model import LinearRegression
import matplotlib.dates as mdates
import pie

file_name = "Copy.xlsx"

data_copy = load_workbook(file_name)
sheet_copy = data_copy.active

data = load_workbook("data.xlsx")
sheet = data.active

added_copy = {}
added = {}
checked = {}
fire = {}

for row in sheet_copy.iter_rows(min_row=2, max_row=sheet_copy.max_row):
    if row[1].value is None:
        break
    added_copy[row[0].value] = row[1].value
    checked[row[0].value] = row[2].value
    fire[row[0].value] = row[3].value

for row in sheet.iter_rows(min_row=2, max_row=sheet_copy.max_row):
    if row[1].value is None:
        break
    added[row[0].value] = row[1].value

checked_list = list(checked.values())
added_copy_list = list(added_copy.values())
added_list = list(added.values())
fire_list = list(fire.values())
calendar = list(checked.keys())

pie.show_pies((fire, "fire"), (checked, "checked"), (added, "added from data.xlxs"), (added_copy, "added actual"))

data_printer = DataHolder(
    calendar,
    file_name,
    (added_copy_list, "Added copy"),
    (added_list, "Added real"),
    (checked_list, "Checked"),
    (fire_list, "Fire")
)

prints_lists_values(data_printer)
try:
    draw_graphics(data_printer, float(input("Write down pause time: ")))
except:
    draw_graphics(data_printer, 0.01)

# new_list = [i * 30 for i in fire_list]
# ratio_list = [new_list[i] / checked_list[i] for i in range(len(checked_list))]
# data_printer_2 = DataHolder(
#     calendar,
#     file_name,
#     (ratio_list, "new_list / checked_list")
# )
# prints_lists_values(data_printer_2)
#
# try:
#     draw_graphics(data_printer_2, float(input("Write down pause time: ")))
# except:
#     draw_graphics(data_printer_2, 0.01)
